# -*- coding: utf-8 -*-
"""
Program to load matlab placenta files into Python + write/read them to pickle files
(Hacky + dependent on current mat data structure..)

Jonathan Lee/Alex Leonard
"""

import scipy as sp, numpy as np, cPickle as pickle
import scipy.io as extracter
import gzip
import glob, os
from openpyxl import load_workbook

from distutils.dir_util import mkpath

path = 'C:\\Users\\ENVY14-i7-SPECTRE\\Documents\\Year 4\AlexMSci\\Placentadirectory\\' 
EARLIxl = path + 'NCS-EARLI BirthWeightData\EARLI specific shapes data.xlsx' #"C:\Users\ENVY14-i7-SPECTRE\Documents\Year 4\AlexMSci\Placentadirectory\NCS-EARLI BirthWeightData\EARLI specific shapes data.xlsx"
NCSxl = path + 'NCS-EARLI BirthWeightData\\NCS specific shapes data.xlsx'
mats = path + 'NCS-EARLI-TWINS ChorionicPlateVascularData\\*\\*.mat'
       
#################################
##### START OF VESSEL CLASS #####
       
class Vessel(object):
    '''Contains a single vessel of the vascular network'''
   
    def __init__(self, terminals, generation, diameter, points, arclength, tortuosity, prefix):
       self.__diameter = diameter
       self.__points = points
       self.__points = self.__points.astype(np.float32)
       self.__arclength = arclength
       self.__generation = generation
       self.__edgePair = terminals
       self.__tortuosity = tortuosity
       
       if prefix=='Z':
           self.__points-=250
       elif prefix=='':
           self.__points*=10
           self.__arclength*=10
           self.__diameter*=10
                 
    def __str__(self):#NOT COMPLETE
        return "Vessel with diameter %s and length %s." % (self.__diameter, self.__arclength)
    def __repr__(self):#NOT COMPLETE
        return "Vessel with diameter %s and length %s." % (self.__diameter, self.__arclength)
       
    def getDiameter(self):
        return self.__diameter
    def getPoints(self):
        return self.__points
    def getArcLength(self):
        return self.__arclength
    def getGeneration(self):
        return self.__generation
    def getEdgePair(self):
        return self.__edgePair
    def getTortuosity(self):
        return self.__tortuosity
        
##### END OF VESSEL CLASS #####              
###############################        
      

##################################
##### START OF NETWORK CLASS #####      
class VascularNetwork(object):
   '''Contains a separate vascular network e.g. veins/arteries with Vessel classes'''
   def __init__(self, skeleton, networkType,prefix):
        self.__points = skeleton[3]
        self.__points=self.__points.astype(np.float32)
        
        if prefix=='Z':
            self.__points-=250

        elif prefix=='':
            self.__points*=10                      
                                
        self.__networkType = networkType
        self.__vessels = []
        self.__angles = []
        self.makeNetwork(skeleton,prefix)
   
   def makeNetwork(self,skeleton,prefix):
       for index,chain in enumerate(skeleton[2][0]):
           temp = Vessel(chain[0][0], chain[1][0][0], chain[2][0][0], chain[4], chain[3][0][0], skeleton[5][0][0][5][index][0],prefix)
           self.__vessels.append(temp)
           
       #Angle pairs written as [(theta1,generation1),(theta2,generation2),....]
       self.__angles=[(x[0],y[0]) for x,y in zip(skeleton[4][0][0][6],skeleton[4][0][0][7])]
               
   def getGraphPoints(self):
       return self.__points

   def getNetworkType(self):
       return self.__networkType
   
   def getVessels(self):
       return self.__vessels
   def getAngles(self):
       return self.__angles
       
   def __str__(self):#NOT COMPLETE
       return self.__networkType+" Network with "+str(len(self.__vessels))+" vessels."
       
   def __repr__(self):#NOT COMPLETE
       return self.__networkType+" Network with "+str(len(self.__vessels))+" vessels."

##### END OF NETWORK CLASS ######              
#################################
           
###################################
##### START OF PLACENTA CLASS #####     
                      
class Placenta(object):
    '''Placenta, units in mm'''
    def __init__(self, ID, group, perimeter, MLP,pfx):
        self.ID = ID
        self.group = group
        self.__perimeter = perimeter
        self.__perimeter=self.__perimeter.astype(np.float32)
        if pfx=='Z':
            self.__perimeter-=250
            self.scan = MLP['scan']
        elif pfx=='':
            self.__perimeter*=10
            self.scan = None
                    
        self.__arteryNetwork = VascularNetwork(MLP[pfx+'skeleton'][0][0],"Artery",pfx)
        self.__veinNetwork = VascularNetwork(MLP[pfx+'skeleton'][0][1],"Vein",pfx)
        
        self.BW = None
        self.GA_wk = None
        self.gender = None
        self.Perimeter = None
        self.Area = None
        self.MaxDiameter = None
        self.OrthDiameter = None
        self.UmbilicDistFromCentroid = None
        self.RadiusMin = None
        self.RadiusMax = None
        self.RadiusMedian = None
        self.Compactness = None
        self.sigma_ins = None
        self.sigma_gc = None
        self.eccentricity_ins = None
        self.eccentricity_gc = None
        self.ThicknessMax_CS = None
        self.ThicknessMean_CS = None
        self.ThicknessStd_CS = None
        
    def __str__(self):#NOT COMPLETE
        return "Placenta %s from group %s" % (self.ID, self.group)
    def __repr__(self):#NOT COMPLETE
        return "Placenta %s from group %s" % (self.ID, self.group)
    
    def getArteryNetwork(self):
        return self.__arteryNetwork
    def getVeinNetwork(self):
        return self.__veinNetwork 
    def getPerimeter(self):
        return self.__perimeter
        
    def setOther(self, weight, gestAge, gender, perim, area, maxdiam, orthdiam, umbdist, rmin, rmax, rmed, compactness, sigma_ins, sigma_gc, ecc_ins, ecc_gc, thkmax, thkmean, thkstd):
        self.BW = weight
        self.GA_wk = gestAge
        self.gender = gender
        self.Perimeter = perim
        self.Area = area
        self.MaxDiameter = maxdiam
        self.OrthDiameter = orthdiam
        self.UmbilicDistFromCentroid = umbdist
        self.RadiusMin = rmin
        self.RadiusMax = rmax
        self.RadiusMedian = rmed
        self.Compactness = compactness
        self.sigma_ins = sigma_ins
        self.sigma_gc = sigma_gc
        self.eccentricity_ins = ecc_ins
        self.eccentricity_gc = ecc_gc
        self.ThicknessMax_CS = thkmax
        self.ThicknessMean_CS = thkmean
        self.ThicknessStd_CS = thkstd
        
##### END OF PLACENTA CLASS #####              
#################################         
        
def loadmats(loc = mats):
    '''Loads files from file directory of matlab .mat files into placentaDict '''
    
    placentaDict = {}
    files = glob.glob(loc)   
     
    for mat in files:
        matlabPlacenta = extracter.loadmat(mat)
        subGroup=os.path.dirname(mat).split('/')[-1]   
        placName=str(os.path.basename(mat).split('.')[0])

        
        #Handle different forms (Twins,NCS,EARLI) from this variable
        placGroup=0
        prefix=""
        if 'TwinsxyONLY' in subGroup:
            placGroup="Twins"
            
        elif 'NCSxyUnScanned' in subGroup:
            placGroup="NCS Unscanned"
            
        elif 'NCSxyxyzScanned' in subGroup:
            placGroup="NCS Scanned"
            placName=placName[1:]
            prefix="Z"
            
        elif 'EARLIxyUnScanned' in subGroup:
            placGroup="EARLI Unscanned"
            placName='E'+placName
            
        elif 'EARLIxyxyzScanned' in subGroup:
            placGroup="EARLI Scanned"
            placName='E'+placName[1:]
            prefix="Z"
        
        print "Adding", placGroup, placName, "from", mat
        
        placentaDict[placName] = \
            Placenta( \
                        ID = placName, \
                        group = placGroup, \
                        perimeter = matlabPlacenta[prefix+'perimeter'], \
                        MLP = matlabPlacenta,\
                        pfx = prefix)
        
    return placentaDict
        
def writePlacentastofiles(placentadict):   
    '''Writes specified placenta dictionary into pickled files'''

         
    mkpath(os.getcwd()+"\data\\")    
    for placenta in placentas.values():
        fileStr=placenta.ID+'.pgz'
        with gzip.open(fileStr, 'wb') as f:
            pickle.dump(placenta, f, -1)

    
#    for placenta in placentadict:
#        print "Writing", placenta
#        pickle.dump(placentadict[placenta], open("data\%s.plc"%placenta, 'wb'), -1)

def loadPlacentasfromdir(loc='H:\Placenta2\Code\Dev1\data\*.pgz'):
    '''Loads and recovers specified pickled placentas into a dictionary'''  
    pyplacentas = {}
    files = glob.glob(loc) 
    
    for x in files: 
        print "Loading file:", x
        with gzip.open(x, 'rb') as f:
            pyplacentas[os.path.basename(x).split(".")[0]] = pickle.load(f)
            #pyplacentas[loaded_object]
        
        #pyplacentas[os.path.basename(x).split(".")[0]] = pickle.load(open(x, 'rb')) #Need ID instead of path x ..
        
    return pyplacentas

def addExcelData(placentadict, excelfile = NCSxl):
    '''Adds data from excel file into placenta dictionary'''  
    newdict = placentadict
    wb = load_workbook(filename = excelfile)
    sheetnames = wb.get_sheet_names()

    for row in wb[sheetnames[0]].rows[1:]:
        if "NCS" in sheetnames[0]:
            tempID="BN"+str(row[1].value)
        elif "EARLI" in sheetnames[0]:
            tempID="E"+str(row[1].value)
        else:
            raise Exception ("Invalid Excel spreadsheet")
            
        try:
            tempPlacenta = newdict[tempID]
            others = [cell.value for cell in row[2:]]
            tempPlacenta.setOther(*others)
            newdict[tempID] = tempPlacenta
            print "Success adding excel data for", newdict[str(tempID)]#, others
       
        except Exception:
            print "Failed adding excel data for " + str(tempID)#, [cell.value for cell in row[2:]]
   
    return newdict
    
def fullData(mats, EARLIxl=EARLIxl, NCSxl=NCSxl):
    '''Produces dictionary of placentas from .mat files with excel data'''  
    
    dict1 = loadmats(mats)    
    print "== ing from EARLI"
    dict2 = addExcelData(dict1, EARLIxl)
    print "== Loading from NCS now"
    dict3 = addExcelData(dict2, NCSxl)
    return dict3

#Test grabbing full data (i.e. loading from mats & then excels)
placentas = fullData(mats)

#Test writing placenta files to data/*.plc
#writePlacentastofiles(placentas, 'test')

#Test deleting placenta dict and loading from data/*.plc
#del placentas
#placentas = loadPlacentasfromdir()
